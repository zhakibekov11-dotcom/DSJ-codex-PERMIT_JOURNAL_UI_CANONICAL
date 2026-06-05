#!/usr/bin/env python3

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def read_json_stdin():
    return json.loads(sys.stdin.buffer.read().decode("utf-8"))


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def validate_payload(payload):
    if not isinstance(payload, dict):
        raise RuntimeError("Передан некорректный payload реестра.")

    request = payload.get("request")
    items = payload.get("items")

    if not isinstance(request, dict):
        raise RuntimeError("Не найдены данные заявки для экспорта реестра.")

    if not isinstance(items, list):
        raise RuntimeError("Не найдены строки заявки для экспорта реестра.")

    return request, items


def autofit(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = normalize(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 48)


def main():
    if len(sys.argv) != 2:
        print("Usage: export_card_request_registry.py <output.xlsx>", file=sys.stderr)
        return 1

    output_path = Path(sys.argv[1])
    payload = read_json_stdin()

    try:
        request, items = validate_payload(payload)
    except Exception as error:  # noqa: BLE001
        print(str(error), file=sys.stderr)
        return 1

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Заявка"
    registry = workbook.create_sheet("Реестр")

    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    header_font = Font(bold=True)

    summary_rows = [
        ("Заявка", request.get("title")),
        ("Тип корочек", request.get("certificateTypeLabel")),
        ("Дата выдачи", request.get("issueDate")),
        ("Создана", request.get("createdAt")),
        ("Создал", request.get("createdByUserName")),
        ("Компания (рус.)", request.get("requestCompanyRu")),
        ("Компания (каз.)", request.get("requestCompanyKz")),
    ]

    for index, (label, value) in enumerate(summary_rows, start=1):
        summary.cell(row=index, column=1, value=label).font = header_font
        summary.cell(row=index, column=2, value=normalize(value))

    registry_headers = [
        "№",
        "Заявка",
        "Тип",
        "ФИО",
        "Должность / квалификация (рус.)",
        "Лауазымы / біліктілік (каз.)",
        "Компания / место работы (рус.)",
        "Компания / место работы (каз.)",
        "Номер удостоверения",
        "Номер протокола",
    ]

    for column_index, title in enumerate(registry_headers, start=1):
        cell = registry.cell(row=1, column=column_index, value=title)
        cell.font = header_font
        cell.fill = header_fill

    for row_index, item in enumerate(items, start=2):
        registry.cell(row=row_index, column=1, value=item.get("index", row_index - 1))
        registry.cell(row=row_index, column=2, value=normalize(request.get("title")))
        registry.cell(row=row_index, column=3, value=normalize(request.get("certificateTypeLabel")))
        registry.cell(row=row_index, column=4, value=normalize(item.get("fullName")))
        registry.cell(row=row_index, column=5, value=normalize(item.get("positionRu")))
        registry.cell(row=row_index, column=6, value=normalize(item.get("positionKz")))
        registry.cell(row=row_index, column=7, value=normalize(item.get("workplaceRu")))
        registry.cell(row=row_index, column=8, value=normalize(item.get("workplaceKz")))
        registry.cell(row=row_index, column=9, value=normalize(item.get("certificateNumber")))
        registry.cell(row=row_index, column=10, value=normalize(item.get("protocolNumber")))

    summary.freeze_panes = "A2"
    registry.freeze_panes = "A2"
    autofit(summary)
    autofit(registry)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
